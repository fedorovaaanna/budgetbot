from pathlib import Path
import aiosqlite
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from config import DB_PATH
from utils import now_local, month_key
from services.cleanup import cleanup_old_transactions
from services.family import get_user_family


async def export_excel(user_id: int):
    await cleanup_old_transactions()
    wb = Workbook()

    ws = wb.active
    ws.title = "Personal operations"
    ws.append(["Дата", "Тип", "Вне бюджета", "Сумма", "Категория", "Комментарий"])

    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("""
        SELECT created_at, type, is_extra, amount, category_name_snapshot, comment
        FROM transactions
        WHERE user_id = ?
        ORDER BY created_at DESC
        """, (user_id,))
        personal_transactions = await cur.fetchall()

        cur = await db.execute("""
        SELECT year, month, income_total, expense_total,
               extra_income_total, extra_expense_total, balance, category_summary_json
        FROM monthly_summaries
        WHERE user_id = ?
        ORDER BY year DESC, month DESC
        """, (user_id,))
        personal_summaries = await cur.fetchall()

    for created_at, type_, is_extra, amount, category, comment in personal_transactions:
        ws.append([created_at, "доход" if type_ == "income" else "расход", "да" if is_extra else "нет", amount, category, comment])

    ws2 = wb.create_sheet("Personal summaries")
    ws2.append(["Месяц", "Доходы", "Расходы", "Внебюджетные доходы", "Внебюджетные расходы", "Итог", "Категории JSON"])
    for year, month, income, expense, extra_income, extra_expense, balance, category_json in personal_summaries:
        ws2.append([month_key(year, month), income, expense, extra_income, extra_expense, balance, category_json])

    family = await get_user_family(user_id)
    if family:
        family_id = family[0]
        ws3 = wb.create_sheet("Family operations")
        ws3.append(["Дата", "Участник", "Тип", "Вне бюджета", "Сумма", "Категория", "Комментарий"])
        async with aiosqlite.connect(DB_PATH) as db:
            cur = await db.execute("""
            SELECT created_at, created_by_member_number, type, is_extra, amount, category_name_snapshot, comment
            FROM family_transactions
            WHERE family_id = ?
            ORDER BY created_at DESC
            """, (family_id,))
            family_transactions = await cur.fetchall()

            cur = await db.execute("""
            SELECT year, month, income_total, expense_total,
                   extra_income_total, extra_expense_total, balance, category_summary_json
            FROM family_monthly_summaries
            WHERE family_id = ?
            ORDER BY year DESC, month DESC
            """, (family_id,))
            family_summaries = await cur.fetchall()

        for created_at, member_number, type_, is_extra, amount, category, comment in family_transactions:
            ws3.append([created_at, f"участник #{member_number}", "доход" if type_ == "income" else "расход", "да" if is_extra else "нет", amount, category, comment])

        ws4 = wb.create_sheet("Family summaries")
        ws4.append(["Месяц", "Доходы", "Расходы", "Внебюджетные доходы", "Внебюджетные расходы", "Итог", "Категории JSON"])
        for year, month, income, expense, extra_income, extra_expense, balance, category_json in family_summaries:
            ws4.append([month_key(year, month), income, expense, extra_income, extra_expense, balance, category_json])

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F2937")
            cell.alignment = Alignment(horizontal="center")
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                max_len = max(max_len, len(str(cell.value or "")))
            sheet.column_dimensions[col_letter].width = min(max_len + 2, 60)

    export_dir = Path("exports")
    export_dir.mkdir(exist_ok=True)
    file_path = export_dir / f"budget_export_user_{user_id}_{now_local().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(file_path)
    return file_path
